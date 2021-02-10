# ! /usr/bin/python
# -*- coding: utf-8 -*-
import sys
import os

import openpyxl
import paralleldots
import time


class Analytics:

    def __init__(self, path, key='juzJfoB9F8wptRaCpX6QFml7IKtvJsdzpHaewDmgrkk'):
        self.path = path
        self.key = key

    def process_file(self):
        if self.exist_file():
            paralleldots.set_api_key(self.key)
            workbook  = openpyxl.load_workbook(self.path)
            sheet = workbook.active
            max_row = sheet.max_row
            sheet.cell(1, 4).value = 'NEGATIVO'
            sheet.cell(1, 5).value = 'NEUTRAL'
            sheet.cell(1, 6).value = 'POSITIVO'
            for row in range(2, 4):
                print(sheet.cell(row, 3).value)
                paralleldots.set_api_key(self.key)
                output_sentiment = paralleldots.sentiment(str(sheet.cell(row, 3).value))
                print(output_sentiment)
                sheet.cell(row, 4).value = round(output_sentiment['sentiment']['negative']*100, 3) if 'sentiment' in output_sentiment else 0
                sheet.cell(row, 5).value = round(output_sentiment['sentiment']['neutral']*100, 3) if 'sentiment' in output_sentiment else 0
                sheet.cell(row, 6).value = round(output_sentiment['sentiment']['positive']*100, 3) if 'sentiment' in output_sentiment else 0
                time.sleep(10)
            workbook.save('Analisis sentimiento procesado.xlsx')
        else:
            print('No existe el archivo para realizar el analisis de información.')

    def exist_file(self):
        return os.path.isfile(self.path)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        print("Parametro -> ",sys.argv[1],'   ',sys.argv)
        analitycs = Analytics(sys.argv[1]) if len(sys.argv) <= 2 else Analytics(sys.argv[2])
        analitycs.process_file()
    else:
        print("Debe seleccionar un archivo.")

